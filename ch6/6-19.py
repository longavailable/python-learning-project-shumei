# -*- coding: utf-8 -*-
"""
Created on Sun Aug 24 12:18:37 2025

@author: Xiaolong Liu
"""

from openpyxl import Workbook

wb = Workbook()

wb.create_sheet('人民币外汇即期报价')
wb.create_sheet('人民币外汇远掉报价')

ws = wb['人民币外汇即期报价']

ws['A1'] = '货币对'

ws['a2'] = '买报价'

ws['b1'] = 'USD/CNY'

ws['b2'] = 6.3644

ws.cell(1, 3, 'EUR/CNY')

ws.cell(2, 3, 6.9210)

ws.append(['卖报价', 6.3655, 6.9210])

wb.save('6-19.xlsx')